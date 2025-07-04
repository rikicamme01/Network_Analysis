import pandas as pd
import sys
import os
sys.path.append(os.path.dirname(sys.path[0]))
from ast import literal_eval
from .utils import clean_text, LABELS
from .models.bert_segmenter import BertSegmenter
from .models.bert_rep import BertRep
from xlsxwriter import Workbook
import pandas as pd
import openpyxl
from openpyxl import load_workbook

class Analyzer():
    def __init__(self) -> None:
        self.bert_seg = BertSegmenter()
        self.bert_rep = BertRep()

    def df_to_excel(self, df, threshold = 50):  # assessment_name x nome file output
        wb = Workbook("Nuovo_file.xlsx", {'nan_inf_to_errors': True})
        ws = wb.add_worksheet("Analisi")

        title_format = wb.add_format({
            'bold': True,
            'font_size': 15,
            'font_color': 'black',
            'bg_color': '#35CE8D',
            'text_wrap': True,
            'center_across': True,
            'border': True
        })
        wrap_format = wb.add_format({
            'text_wrap': True,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })
        default_format = wb.add_format({
            'text_wrap': False,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })
        highlight_format = wb.add_format({
            'bg_color': '#fdff32',
            'text_wrap': True,
            'border': True,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })

        # Imposta i titoli delle colonne
        for i, col in enumerate(df.columns):
            if col in ['età', 'Età', 'age', 'Age', 'index']:
                ws.set_column(i, i, 6)
            elif col == 'Repertorio':
                ws.write_column('Z1', LABELS)
                ws.data_validation(1, i, 1048575, i, {'validate': 'list', 'source': '$Z$1:$K$24'})
                ws.set_column(i, i, 25)
            elif col in ['Genere', 'Ruolo']:
                ws.set_column(i, i, 15)
            else:
                ws.set_column(i, i, 30)

            ws.write(0, i, col, title_format)

        index_str = 1
        index_alt = 1
        index_rep = 1
        highlight = False

        for i_row, row in df.iterrows():
            # Variabile per tracciare se è il primo stralcio
            prev_row = None

            for col_name, item in row.items():
                if col_name == 'Stralcio':
                    for i_stralci, stralci in enumerate(item):
                        # Copia la riga precedente (prev_row) per tutte le colonne tranne 'Stralcio'
                        if prev_row is not None:
                            for i_col, col_val in enumerate(prev_row):
                                if df.columns[i_col] != 'Stralcio':  # Non sovrascrivere la colonna 'Stralcio'
                                    ws.write(index_str, i_col, col_val, wrap_format)

                        # Scrivi lo stralcio
                        ws.write(index_str, df.columns.get_loc(col_name), stralci, wrap_format)
                        index_str += 1  # Incrementa solo quando scrivi uno stralcio

                    prev_row = row  # Memorizza la riga corrente per copiarla nel prossimo stralcio

                elif col_name == 'Alternative':  # Simile a Stralcio, ma con logica diversa
                    for i_dict, dictio in enumerate(item):
                        dict_str = ''
                        for key, rep in dictio.items():
                            dict_str += f"{key}: {rep}%\n"
                            if list(dictio.keys()).index(key) == 0 and rep <= threshold:
                                highlight = True
                        dict_str = dict_str.rstrip("\n")
                        ws.write(index_alt, df.columns.get_loc(col_name), dict_str, wrap_format)
                        if highlight:
                            ws.write(index_alt, df.columns.get_loc(col_name), dict_str, highlight_format)

                        index_alt += 1
                        highlight = False

                elif col_name == 'Repertorio':
                    for i_first_rep, first_rep in enumerate(item):
                        ws.write(index_rep, df.columns.get_loc(col_name), first_rep, wrap_format)
                        index_rep += 1

                else:
                    ws.write(index_str, df.columns.get_loc(col_name), item, default_format)

            index_str += 1  # Incrementa index_str solo alla fine della riga completa

        wb.close()
        return wb

    def predict(self, df):
        if not isinstance(df, pd.DataFrame):
            raise Exception('L\'oggetto fornito non è un DataFrame valido.')
        
        # Log per verificare il DataFrame iniziale
        print("DataFrame iniziale:", df.head())
        
        # Controlla che il DataFrame contenga la colonna 'Testo'
        if 'Testo' in df.columns:
            print("Colonna 'Testo' trovata.")
            df['Testo'] = df['Testo'].map(clean_text)
            # Log per vedere il testo pulito
            print("Testo dopo la pulizia:", df['Testo'].head())
        else:
            raise Exception('The uploaded DataFrame is missing the "Testo" column')
        
        # Predizione per la colonna 'Stralcio'
        try:
            df['Stralcio'] = df['Testo'].map(self.bert_seg.predict).values.tolist()
            # Log per verificare la colonna 'Stralcio'
            print("Stralcio dopo la predizione:", df['Stralcio'].head())
        except Exception as e:
            print(f"Errore durante la predizione per 'Stralcio': {str(e)}")
            raise

        # Creazione delle colonne 'Repertorio' e 'Alternative'
        list_column_alt = []
        list_column_rep = []
        for i, str_list in df['Stralcio'].items():
            try:
                list_dict = self.bert_rep.predict_vector(str_list)
                # Log per vedere il dizionario restituito da predict_vector
                print(f"Predizioni per {i}: {list_dict}")
            except Exception as e:
                print(f"Errore durante la predizione per 'Repertorio' e 'Alternative' alla riga {i}: {str(e)}")
                continue

            list_dict_sorted = []
            list_first_rep = []
            for dizio in list_dict:
                # Ordina il dizionario e seleziona i primi 5
                sorted_dict = sorted(dizio.items(), key=lambda item: item[1], reverse=True)
                dizio = dict(sorted_dict[:5])
                
                # Rounding dei valori in percentuale
                for key, value in dizio.items():
                    try:
                        dizio[key] = round(value * 100, 2 - int(f"{value * 100:.0e}".split('e')[1]))
                    except Exception as e:
                        print(f"Errore nel rounding per {key}: {str(e)}")
                        continue

                list_dict_sorted.append(dizio)
                list_first_rep.append(next(iter(dizio)))  # Primo elemento del dizionario

            list_column_alt.append(list_dict_sorted)
            list_column_rep.append(list_first_rep)
        
        # Log per vedere le colonne 'Repertorio' e 'Alternative' prima di aggiungerle
        print("Repertorio prima di aggiungere:", list_column_rep[:5])
        print("Alternative prima di aggiungere:", list_column_alt[:5])

        df['Repertorio'] = list_column_rep
        df['Alternative'] = list_column_alt

        # Se non esiste la colonna 'Ads', aggiungila come colonna vuota
        if 'Ads' not in df.columns:
            df['Ads'] = ''
            print("Colonna 'Ads' aggiunta.")
        
        # Log per vedere il DataFrame finale
        print("DataFrame finale:", df.head())

        # Restituisci il DataFrame modificato
        return df


####################################################################################################################
'''
    import pandas as pd
import sys
import os
sys.path.append(os.path.dirname(sys.path[0]))
from ast import literal_eval
from utils import find_word_bounds, clean_text
from models.bert_segmenter import BertSegmenter
from models.bert_rep import BertRep
from utils import LABELS 
from xlsxwriter import Workbook

class Analyzer():
    def __init__(self) -> None:
        self.bert_seg = BertSegmenter()
        self.bert_rep = BertRep()

    def df_to_excel(self, df, threshold):  # assessment_name x nome file output
        wb = Workbook("Nuovo_file.xlsx", {'nan_inf_to_errors': True})
        ws = wb.add_worksheet("Analisi")

        title_format = wb.add_format({
            'bold': True,
            'font_size': 15,
            'font_color': 'black',
            'bg_color': '#35CE8D',
            'text_wrap': True,
            'center_across': True,
            'border': True
        })
        wrap_format = wb.add_format({
            'text_wrap': True,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })
        default_format = wb.add_format({
            'text_wrap': False,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })
        highlight_format = wb.add_format({
            'bg_color': '#fdff32',
            'text_wrap': True,
            'border': True,
            'font_size': 15,
            'align': 'left',
            'align': 'top',
        })

        # Imposta i titoli delle colonne
        for i, col in enumerate(df.columns):
            if col in ['età', 'Età', 'age', 'Age', 'index']:
                ws.set_column(i, i, 6)
            elif col == 'Repertorio':
                ws.write_column('Z1', LABELS)
                ws.data_validation(1, i, 1048575, i, {'validate': 'list', 'source': '$Z$1:$K$24'})
                ws.set_column(i, i, 25)
            elif col in ['Genere', 'Ruolo']:
                ws.set_column(i, i, 15)
            else:
                ws.set_column(i, i, 30)

            ws.write(0, i, col, title_format)

        index_str = 1
        index_alt = 1
        index_rep = 1
        highlight = False

        for i_row, row in df.iterrows():
            for col_name, item in row.items():
                if col_name == 'Stralcio':
                    for i_stralci, stralci in enumerate(item):
                        if i_stralci > 0:  # Per evitare di creare una nuova riga vuota
                            # Copia la riga precedente (index_str - 1) per tutte le colonne
                            for i_col, col_val in enumerate(row):
                                if col_name != df.columns[i_col]:  # Non sovrascrivere la colonna stralcio
                                    ws.write(index_str, i_col, col_val, wrap_format)
                        
                        ws.write(index_str, df.columns.get_loc(col_name), stralci, wrap_format)
                        index_str += 1  # Aggiorna solo quando scrivi un stralcio

                elif col_name == 'Alternative':  # Simile a Stralcio, ma con logica diversa
                    for i_dict, dictio in enumerate(item):
                        dict_str = ''
                        for key, rep in dictio.items():
                            dict_str += f"{key}: {rep}%\n"
                            if list(dictio.keys()).index(key) == 0 and rep <= threshold:
                                highlight = True
                        dict_str = dict_str.rstrip("\n")
                        ws.write(index_alt, df.columns.get_loc(col_name), dict_str, wrap_format)
                        if highlight:
                            ws.write(index_alt, df.columns.get_loc(col_name), dict_str, highlight_format)

                        index_alt += 1
                        highlight = False

                elif col_name == 'Repertorio':
                    for i_first_rep, first_rep in enumerate(item):
                        ws.write(index_rep, df.columns.get_loc(col_name), first_rep, wrap_format)
                        index_rep += 1

                else:
                    ws.write(index_str, df.columns.get_loc(col_name), item, default_format)

            index_str += 1  # Incrementa index_str solo alla fine della riga completa

        wb.close()
        return wb


    
    def predict(self, path_file):
        threshold = 50
        extension = os.path.splitext(path_file)[1]

        if extension == '.xlsx':
            df = pd.read_excel(path_file,sheet_name=1, converters={'Stralci': literal_eval, 'Repertori': literal_eval} )
        elif extension == '.csv':
            df = pd.read_csv(path_file, converters={'Stralci': literal_eval, 'Repertori': literal_eval})
        elif isinstance(path_file, pd.DataFrame):
            df= path_file
        else:
            raise Exception('Extension of file not supported')
        
        #carica csv
        if 'Testo' in df.columns:
            df['Testo'] = df['Testo'].map(clean_text)
            #df['Stralci'] = df['Stralci'].map(lambda x: [clean_text(s) for s in x])
            #df['Bounds'] = df.apply(lambda x: find_word_bounds(x['Stralci'], x['Testo']), axis=1).values.tolist()
        else:
            raise Exception('The uploaded file is missing the "Testo" column')

        #predict
        df['Stralcio'] = df['Testo'].map(self.bert_seg.predict).values.tolist()
        #df['Bounds_predetti'] = df.apply(lambda x: find_word_bounds(x['Stralci_predetti'], x['Testo']), axis=1).values.tolist()

        list_column_alt = []
        list_column_rep = []
        for i, str_list in df['Stralcio'].items(): #str_list is a list of stralci related to one text
            list_dict = self.bert_rep.predict_vector(str_list)
            list_dict_sorted = []
            list_first_rep = []
            for dizio in list_dict:
                sorted_dict = sorted(dizio.items(), key=lambda item: item[1], reverse = True)
                dizio = dict(sorted_dict[:5])
                for key, value in dizio.items():
                    dizio[key] = round(value * 100, 2 - int(f"{value * 100:.0e}".split('e')[1]))

                list_dict_sorted.append(dizio)
                list_first_rep.append(next(iter(dizio)))
             
            list_column_alt.append(list_dict_sorted)
            list_column_rep.append(list_first_rep)
             
            #eventuale aggiornamento contatore x progress bar
            str_column_len = len(df['Stralcio'])
            print(f'Testo numero {i+1}/{ str_column_len} denominato')

        df['Repertorio'] = list_column_rep
        df['Alternative'] = list_column_alt

        if 'Ads' not in df.columns:
            df['Ads'] = ''
        
        # Ritorna al file Excel esistente e salva le modifiche
        with pd.ExcelWriter(path_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Salva il DataFrame modificato nel secondo foglio
            df.to_excel(writer, index=False, sheet_name='Analisi', startrow=0)

        workbook = self.df_to_excel(df, threshold)

        #salvataggio/esporta file 
    
    def progress():
        pass

'''
